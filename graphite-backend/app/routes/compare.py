from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from app.models.experiment import (
    Experiment, ExperimentBasic, ExperimentPi, ExperimentLoose,
    ExperimentCarbon, ExperimentGraphite, ExperimentRolling, 
    ExperimentProduct
)
from app.models.user import User
from app import db
import traceback
from decimal import Decimal
from datetime import datetime, date

# ============ Excel导出相关导入 ============
import io
from urllib.parse import quote
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

compare_bp = Blueprint('compare', __name__)

# ==========================================
# 🔧 SQLAlchemy对象序列化辅助函数（修复Decimal小数位问题）
# ==========================================
def model_to_dict(instance):
    """
    将SQLAlchemy模型对象转换为字典
    
    修复: Decimal类型会保留所有小数位（包括尾随零），导致显示999.000000
    解决: 转换为float后去除尾随零
    
    Args:
        instance: SQLAlchemy模型实例
        
    Returns:
        dict: 包含所有列数据的字典
    """
    if instance is None:
        return {}
    
    result = {}
    # 遍历所有列
    for column in instance.__table__.columns:
        value = getattr(instance, column.name)
        
        # 🔧 修复：处理Decimal类型，去除尾随零
        if isinstance(value, Decimal):
            # 转换为float（去除尾随零）
            # 如果是整数（如999.000000），会变成999.0
            # 如果是小数（如0.850000），会变成0.85
            float_value = float(value)
            # Python的float会自动去除尾随零
            result[column.name] = float_value
        # 处理日期时间类型
        elif hasattr(value, 'isoformat'):
            result[column.name] = value.isoformat()
        else:
            result[column.name] = value
    
    return result


@compare_bp.route('/compare', methods=['POST'])
@jwt_required()
def compare_experiments():
    """
    实验数据对比API
    
    请求体：
    {
        "experiment_ids": [1, 2, 3, ...]  # 实验ID列表（2-10个）
    }
    
    返回：
    {
        "experiments": [
            {
                "id": 1,
                "code": "100ISA-TH5100-251008DG-RIF01",
                "basic": {...},
                "pi": {...},
                "carbon": {...},
                "graphite": {...},
                "product": {...}
            },
            ...
        ],
        "fields": [
            {
                "category": "基本参数",
                "name": "PI膜厚度",
                "key": "pi_film_thickness",
                "type": "number",
                "unit": "μm"
            },
            ...
        ]
    }
    """
    try:
        print("=" * 60)
        print("📊 收到实验对比请求")
        
        # 1. 验证JWT
        current_user_id = int(get_jwt_identity())
        user = User.query.get(current_user_id)
        print(f"   - 用户ID: {current_user_id}, 角色: {user.role}")
        
        # 2. 权限检查（只有工程师和管理员可以对比）
        if user.role not in ['admin', 'engineer']:
            print(f"❌ 权限不足: {user.role}")
            return jsonify({'error': '您没有权限访问此功能'}), 403
        
        # 3. 获取请求参数
        data = request.get_json()
        experiment_ids = data.get('experiment_ids', [])
        print(f"   - 请求对比的实验ID: {experiment_ids}")
        
        # 4. 验证实验数量（2-10个）
        if len(experiment_ids) < 2:
            print("❌ 实验数量不足")
            return jsonify({'error': '请至少选择2个实验进行对比'}), 400
        if len(experiment_ids) > 10:
            print("❌ 实验数量过多")
            return jsonify({'error': '最多只能同时对比10个实验'}), 400
        
        # 5. 查询实验数据
        experiments_data = []
        for exp_id in experiment_ids:
            experiment = Experiment.query.get(exp_id)
            if not experiment:
                print(f"❌ 实验ID {exp_id} 不存在")
                return jsonify({'error': f'实验ID {exp_id} 不存在'}), 404
            
            # 查询所有子表数据
            basic = ExperimentBasic.query.filter_by(experiment_id=exp_id).first()
            pi = ExperimentPi.query.filter_by(experiment_id=exp_id).first()
            loose = ExperimentLoose.query.filter_by(experiment_id=exp_id).first()
            carbon = ExperimentCarbon.query.filter_by(experiment_id=exp_id).first()
            graphite = ExperimentGraphite.query.filter_by(experiment_id=exp_id).first()
            rolling = ExperimentRolling.query.filter_by(experiment_id=exp_id).first()
            product = ExperimentProduct.query.filter_by(experiment_id=exp_id).first()
            
            # 🔧 使用修正后的 model_to_dict()（自动去除Decimal尾随零）
            exp_data = {
                'id': experiment.id,
                'code': experiment.experiment_code,
                'status': experiment.status,
                'created_at': experiment.created_at.isoformat() if experiment.created_at else None,
                'basic': model_to_dict(basic),
                'pi': model_to_dict(pi),
                'loose': model_to_dict(loose),
                'carbon': model_to_dict(carbon),
                'graphite': model_to_dict(graphite),
                'rolling': model_to_dict(rolling),
                'product': model_to_dict(product)
            }
            experiments_data.append(exp_data)
            print(f"   ✅ 加载实验 {experiment.experiment_code}")
        
        # 6. 定义字段元数据（所有要对比的字段）
        fields = _get_comparison_fields()
        
        print(f"✅ 对比成功: {len(experiments_data)}个实验, {len(fields)}个字段")
        print("=" * 60)
        
        return jsonify({
            'experiments': experiments_data,
            'fields': fields
        }), 200
        
    except Exception as e:
        print(f"❌ 对比实验数据失败: {str(e)}")
        traceback.print_exc()
        print("=" * 60)
        return jsonify({'error': '对比实验数据失败'}), 500


def _get_comparison_fields():
    """
    定义所有要对比的字段（方案B：完整Schema覆盖）
    
    更新日期: 2024-12-09
    更新内容:
    1. 完整覆盖所有Schema字段（95个）
    2. 排除9个照片/文件字段
    3. 保留所有文本说明字段（有对比价值）
    4. 保持简称不转换
    
    总计: 86个对比字段
    
    返回字段元数据列表，包含：
    - category: 分类（基本参数、PI膜参数等）
    - name: 中文名称
    - key: 字段key（用于取值，支持嵌套如 basic.pi_film_thickness）
    - type: 数据类型（number/string/date/datetime）
    - unit: 单位
    """
    return [
        # ============ 基本参数 (12个字段) ============
        {"category": "基本参数", "name": "实验编码", "key": "code", "type": "string", "unit": ""},
        {"category": "基本参数", "name": "实验日期", "key": "basic.experiment_date", "type": "date", "unit": ""},
        {"category": "基本参数", "name": "PI膜厚度", "key": "basic.pi_film_thickness", "type": "number", "unit": "μm"},
        {"category": "基本参数", "name": "客户类型", "key": "basic.customer_type", "type": "string", "unit": ""},
        {"category": "基本参数", "name": "客户名称", "key": "basic.customer_name", "type": "string", "unit": ""},
        {"category": "基本参数", "name": "PI膜型号", "key": "basic.pi_film_model", "type": "string", "unit": ""},
        {"category": "基本参数", "name": "烧制地点", "key": "basic.sintering_location", "type": "string", "unit": ""},
        {"category": "基本参数", "name": "石墨型号", "key": "basic.graphite_model", "type": "string", "unit": ""},
        {"category": "基本参数", "name": "送烧材料类型", "key": "basic.material_type_for_firing", "type": "string", "unit": ""},
        {"category": "基本参数", "name": "压延方式", "key": "basic.rolling_method", "type": "string", "unit": ""},
        {"category": "基本参数", "name": "实验编组", "key": "basic.experiment_group", "type": "number", "unit": ""},
        {"category": "基本参数", "name": "实验目的", "key": "basic.experiment_purpose", "type": "string", "unit": ""},
        
        # ============ PI膜参数 (8个字段) ============
        {"category": "PI膜参数", "name": "PI膜厂商", "key": "pi.pi_manufacturer", "type": "string", "unit": ""},
        {"category": "PI膜参数", "name": "PI膜厚度", "key": "pi.pi_thickness_detail", "type": "number", "unit": "μm"},
        {"category": "PI膜参数", "name": "PI膜型号详细", "key": "pi.pi_model_detail", "type": "string", "unit": ""},
        {"category": "PI膜参数", "name": "PI膜宽幅", "key": "pi.pi_width", "type": "number", "unit": "mm"},
        {"category": "PI膜参数", "name": "PI支料号/批次号", "key": "pi.pi_roll_batch_number", "type": "string", "unit": ""},
        {"category": "PI膜参数", "name": "PI重量", "key": "pi.pi_weight", "type": "number", "unit": "kg"},
        {"category": "PI膜参数", "name": "烧制卷数", "key": "pi.firing_rolls", "type": "number", "unit": "卷"},
        {"category": "PI膜参数", "name": "PI膜补充说明", "key": "pi.pi_notes", "type": "string", "unit": ""},
        
        # ============ 松卷参数 (4个字段) ============
        {"category": "松卷参数", "name": "卷芯筒类型", "key": "loose.core_tube_type", "type": "string", "unit": ""},
        {"category": "松卷参数", "name": "松卷间隙(卷内)", "key": "loose.loose_gap_inner", "type": "number", "unit": "μm"},
        {"category": "松卷参数", "name": "松卷间隙(卷中)", "key": "loose.loose_gap_middle", "type": "number", "unit": "μm"},
        {"category": "松卷参数", "name": "松卷间隙(卷外)", "key": "loose.loose_gap_outer", "type": "number", "unit": "μm"},
        
        # ============ 碳化参数 (18个字段，排除3个照片/文件) ============
        {"category": "碳化参数", "name": "碳化炉编号", "key": "carbon.carbon_furnace_number", "type": "string", "unit": ""},
        {"category": "碳化参数", "name": "碳化炉次", "key": "carbon.carbon_furnace_batch", "type": "number", "unit": ""},
        {"category": "碳化参数", "name": "舟皿型号", "key": "carbon.boat_model", "type": "string", "unit": ""},
        {"category": "碳化参数", "name": "包裹形式", "key": "carbon.wrapping_method", "type": "string", "unit": ""},
        {"category": "碳化参数", "name": "真空度", "key": "carbon.vacuum_degree", "type": "number", "unit": ""},
        {"category": "碳化参数", "name": "电量", "key": "carbon.power_consumption", "type": "number", "unit": ""},
        {"category": "碳化参数", "name": "开机时间", "key": "carbon.start_time", "type": "datetime", "unit": ""},
        {"category": "碳化参数", "name": "关机时间", "key": "carbon.end_time", "type": "datetime", "unit": ""},
        {"category": "碳化参数", "name": "碳化温度1", "key": "carbon.carbon_temp1", "type": "number", "unit": "℃"},
        {"category": "碳化参数", "name": "碳化厚度1", "key": "carbon.carbon_thickness1", "type": "number", "unit": "μm"},
        {"category": "碳化参数", "name": "碳化温度2", "key": "carbon.carbon_temp2", "type": "number", "unit": "℃"},
        {"category": "碳化参数", "name": "碳化厚度2", "key": "carbon.carbon_thickness2", "type": "number", "unit": "μm"},
        {"category": "碳化参数", "name": "碳化最高温度", "key": "carbon.carbon_max_temp", "type": "number", "unit": "℃"},
        {"category": "碳化参数", "name": "碳化总时长", "key": "carbon.carbon_total_time", "type": "number", "unit": "min"},
        {"category": "碳化参数", "name": "碳化后厚度", "key": "carbon.carbon_film_thickness", "type": "number", "unit": "μm"},
        {"category": "碳化参数", "name": "碳化后重量", "key": "carbon.carbon_after_weight", "type": "number", "unit": "kg"},
        {"category": "碳化参数", "name": "成碳率", "key": "carbon.carbon_yield_rate", "type": "number", "unit": "%"},
        {"category": "碳化参数", "name": "碳化补充说明", "key": "carbon.carbon_notes", "type": "string", "unit": ""},
        # ❌ 排除: carbon_loading_photo, carbon_sample_photo, carbon_other_params
        
        # ============ 石墨化参数 (25个字段，排除3个照片/文件) ============
        {"category": "石墨化参数", "name": "石墨化炉编号", "key": "graphite.graphite_furnace_number", "type": "string", "unit": ""},
        {"category": "石墨化参数", "name": "石墨化炉次", "key": "graphite.graphite_furnace_batch", "type": "number", "unit": ""},
        {"category": "石墨化参数", "name": "开机时间", "key": "graphite.graphite_start_time", "type": "datetime", "unit": ""},
        {"category": "石墨化参数", "name": "关机时间", "key": "graphite.graphite_end_time", "type": "datetime", "unit": ""},
        {"category": "石墨化参数", "name": "气压值", "key": "graphite.gas_pressure", "type": "number", "unit": ""},
        {"category": "石墨化参数", "name": "电量", "key": "graphite.graphite_power", "type": "number", "unit": ""},
        {"category": "石墨化参数", "name": "石墨化温度1", "key": "graphite.graphite_temp1", "type": "number", "unit": "℃"},
        {"category": "石墨化参数", "name": "石墨化厚度1", "key": "graphite.graphite_thickness1", "type": "number", "unit": "μm"},
        {"category": "石墨化参数", "name": "石墨化温度2", "key": "graphite.graphite_temp2", "type": "number", "unit": "℃"},
        {"category": "石墨化参数", "name": "石墨化厚度2", "key": "graphite.graphite_thickness2", "type": "number", "unit": "μm"},
        {"category": "石墨化参数", "name": "石墨化温度3", "key": "graphite.graphite_temp3", "type": "number", "unit": "℃"},
        {"category": "石墨化参数", "name": "石墨化厚度3", "key": "graphite.graphite_thickness3", "type": "number", "unit": "μm"},
        {"category": "石墨化参数", "name": "石墨化温度4", "key": "graphite.graphite_temp4", "type": "number", "unit": "℃"},
        {"category": "石墨化参数", "name": "石墨化厚度4", "key": "graphite.graphite_thickness4", "type": "number", "unit": "μm"},
        {"category": "石墨化参数", "name": "石墨化温度5", "key": "graphite.graphite_temp5", "type": "number", "unit": "℃"},
        {"category": "石墨化参数", "name": "石墨化厚度5", "key": "graphite.graphite_thickness5", "type": "number", "unit": "μm"},
        {"category": "石墨化参数", "name": "石墨化温度6", "key": "graphite.graphite_temp6", "type": "number", "unit": "℃"},
        {"category": "石墨化参数", "name": "石墨化厚度6", "key": "graphite.graphite_thickness6", "type": "number", "unit": "μm"},
        {"category": "石墨化参数", "name": "卷内发泡厚度", "key": "graphite.inner_foaming_thickness", "type": "number", "unit": "μm"},
        {"category": "石墨化参数", "name": "卷外发泡厚度", "key": "graphite.outer_foaming_thickness", "type": "number", "unit": "μm"},
        {"category": "石墨化参数", "name": "石墨化最高温度", "key": "graphite.graphite_max_temp", "type": "number", "unit": "℃"},
        {"category": "石墨化参数", "name": "石墨宽幅", "key": "graphite.graphite_width", "type": "number", "unit": "mm"},
        {"category": "石墨化参数", "name": "收缩比", "key": "graphite.shrinkage_ratio", "type": "number", "unit": "%"},
        {"category": "石墨化参数", "name": "石墨化总时长", "key": "graphite.graphite_total_time", "type": "number", "unit": "min"},
        {"category": "石墨化参数", "name": "石墨化后重量", "key": "graphite.graphite_after_weight", "type": "number", "unit": "kg"},
        {"category": "石墨化参数", "name": "石墨化收率", "key": "graphite.graphite_yield_rate", "type": "number", "unit": "%"},
        {"category": "石墨化参数", "name": "石墨最薄极限", "key": "graphite.graphite_min_thickness", "type": "number", "unit": "μm"},
        {"category": "石墨化参数", "name": "石墨化补充说明", "key": "graphite.graphite_notes", "type": "string", "unit": ""},
        # ❌ 排除: graphite_loading_photo, graphite_sample_photo, graphite_other_params
        
        # ============ 压延参数 (5个字段) ============
        {"category": "压延参数", "name": "压延机台", "key": "rolling.rolling_machine", "type": "string", "unit": ""},
        {"category": "压延参数", "name": "压延压力", "key": "rolling.rolling_pressure", "type": "number", "unit": "MPa"},
        {"category": "压延参数", "name": "压延张力", "key": "rolling.rolling_tension", "type": "number", "unit": ""},
        {"category": "压延参数", "name": "压延速度", "key": "rolling.rolling_speed", "type": "number", "unit": "m/s"},
        {"category": "压延参数", "name": "压延补充说明", "key": "rolling.rolling_notes", "type": "string", "unit": ""},
        
        # ============ 成品参数 (14个字段，排除3个照片/文件) ============
        {"category": "成品参数", "name": "成品编码", "key": "product.product_code", "type": "string", "unit": ""},
        {"category": "成品参数", "name": "样品平均厚度", "key": "product.avg_thickness", "type": "number", "unit": "μm"},
        {"category": "成品参数", "name": "规格", "key": "product.specification", "type": "string", "unit": ""},
        {"category": "成品参数", "name": "平均密度", "key": "product.avg_density", "type": "number", "unit": "g/cm³"},
        {"category": "成品参数", "name": "热扩散系数", "key": "product.thermal_diffusivity", "type": "number", "unit": "mm²/s"},
        {"category": "成品参数", "name": "导热系数", "key": "product.thermal_conductivity", "type": "number", "unit": "W/m·K"},
        {"category": "成品参数", "name": "比热", "key": "product.specific_heat", "type": "number", "unit": "J/g/K"},
        {"category": "成品参数", "name": "内聚力", "key": "product.cohesion", "type": "number", "unit": "gf"},
        {"category": "成品参数", "name": "剥离力", "key": "product.peel_strength", "type": "number", "unit": "gf"},
        {"category": "成品参数", "name": "粗糙度", "key": "product.roughness", "type": "string", "unit": ""},
        {"category": "成品参数", "name": "外观及不良描述", "key": "product.appearance_desc", "type": "string", "unit": ""},
        {"category": "成品参数", "name": "实验总结", "key": "product.experiment_summary", "type": "string", "unit": ""},
        {"category": "成品参数", "name": "备注", "key": "product.remarks", "type": "string", "unit": ""},
        {"category": "成品参数", "name": "结合力", "key": "product.bond_strength", "type": "number", "unit": ""},
        # ❌ 排除: appearance_defect_photo, sample_photo, other_files
    ]


# ============================================================================
# Excel导出功能
# ============================================================================

# Excel导出字段顺序定义（严格控制导出顺序）
EXPORT_FIELD_ORDER = [
    # 基本信息
    {'key': 'code', 'name': '实验编码', 'format': None},
    {'key': 'created_at', 'name': '实验日期', 'format': None},
    
    # 基本参数（10个）
    {'key': 'basic.pi_film_thickness', 'name': 'PI膜厚度', 'format': '0.00', 'unit': 'μm'},
    {'key': 'basic.customer_type', 'name': '客户类型', 'format': None},
    {'key': 'basic.customer_name', 'name': '客户名称', 'format': None},
    {'key': 'basic.pi_film_model', 'name': 'PI膜型号', 'format': None},
    {'key': 'basic.sintering_location', 'name': '烧制地点', 'format': None},
    {'key': 'basic.graphite_model', 'name': '石墨型号', 'format': None},
    {'key': 'basic.material_type_for_firing', 'name': '送烧材料类型', 'format': None},
    {'key': 'basic.rolling_method', 'name': '压延方式', 'format': None},
    {'key': 'basic.experiment_group', 'name': '实验编组', 'format': '0'},
    {'key': 'basic.experiment_purpose', 'name': '实验目的', 'format': None},  # ✅ 新增
    
    # PI膜参数（8个）
    {'key': 'pi.pi_manufacturer', 'name': 'PI膜厂商', 'format': None},
    {'key': 'pi.pi_thickness_detail', 'name': 'PI膜详细厚度', 'format': '0.00', 'unit': 'μm'},
    {'key': 'pi.pi_model_detail', 'name': 'PI膜详细型号', 'format': None},
    {'key': 'pi.pi_width', 'name': 'PI膜宽幅', 'format': '0.00', 'unit': 'mm'},
    {'key': 'pi.pi_roll_batch_number', 'name': 'PI支料号/批次号', 'format': None},
    {'key': 'pi.pi_weight', 'name': 'PI重量', 'format': '0.000', 'unit': 'kg'},
    {'key': 'pi.firing_rolls', 'name': '烧制卷数', 'format': '0'},
    {'key': 'pi.pi_notes', 'name': 'PI膜补充说明', 'format': None},  # ✅ 新增
    
    # 松卷参数（3个）
    {'key': 'loose.loose_gap_inner', 'name': '松卷间隙卷内', 'format': '0.00', 'unit': 'μm'},
    {'key': 'loose.loose_gap_middle', 'name': '松卷间隙卷中', 'format': '0.00', 'unit': 'μm'},
    {'key': 'loose.loose_gap_outer', 'name': '松卷间隙卷外', 'format': '0.00', 'unit': 'μm'},
    
    # 碳化参数（14个）
    {'key': 'carbon.carbon_furnace_number', 'name': '碳化炉编号', 'format': None},
    {'key': 'carbon.carbon_furnace_batch', 'name': '碳化炉次', 'format': '0'},
    {'key': 'carbon.carbon_temp1', 'name': '碳化温度1', 'format': '0', 'unit': '℃'},
    {'key': 'carbon.carbon_thickness1', 'name': '碳化厚度1', 'format': '0.00', 'unit': 'μm'},
    {'key': 'carbon.carbon_temp2', 'name': '碳化温度2', 'format': '0', 'unit': '℃'},
    {'key': 'carbon.carbon_thickness2', 'name': '碳化厚度2', 'format': '0.00', 'unit': 'μm'},
    {'key': 'carbon.carbon_max_temp', 'name': '碳化最高温度', 'format': '0.00', 'unit': '℃'},
    {'key': 'carbon.carbon_total_time', 'name': '碳化总时长', 'format': '0', 'unit': 'min'},
    {'key': 'carbon.carbon_film_thickness', 'name': '碳化膜厚度', 'format': '0.00', 'unit': 'μm'},
    {'key': 'carbon.carbon_after_weight', 'name': '碳化后重量', 'format': '0.000', 'unit': 'kg'},
    {'key': 'carbon.carbon_yield_rate', 'name': '碳化成碳率', 'format': '0.00', 'unit': '%'},
    {'key': 'carbon.vacuum_degree', 'name': '真空度', 'format': '0.0000'},
    {'key': 'carbon.power_consumption', 'name': '碳化电量', 'format': '0.00'},
    {'key': 'carbon.carbon_notes', 'name': '碳化补充说明', 'format': None},  # ✅ 新增
    
    # 石墨化参数（25个）
    {'key': 'graphite.graphite_furnace_number', 'name': '石墨炉编号', 'format': None},
    {'key': 'graphite.graphite_furnace_batch', 'name': '石墨化炉次', 'format': '0'},
    {'key': 'graphite.graphite_temp1', 'name': '石墨化温度1', 'format': '0.00', 'unit': '℃'},
    {'key': 'graphite.graphite_thickness1', 'name': '石墨化厚度1', 'format': '0.00', 'unit': 'μm'},
    {'key': 'graphite.graphite_temp2', 'name': '石墨化温度2', 'format': '0.00', 'unit': '℃'},
    {'key': 'graphite.graphite_thickness2', 'name': '石墨化厚度2', 'format': '0.00', 'unit': 'μm'},
    {'key': 'graphite.graphite_temp3', 'name': '石墨化温度3', 'format': '0.00', 'unit': '℃'},
    {'key': 'graphite.graphite_thickness3', 'name': '石墨化厚度3', 'format': '0.00', 'unit': 'μm'},
    {'key': 'graphite.graphite_temp4', 'name': '石墨化温度4', 'format': '0.00', 'unit': '℃'},
    {'key': 'graphite.graphite_thickness4', 'name': '石墨化厚度4', 'format': '0.00', 'unit': 'μm'},
    {'key': 'graphite.graphite_temp5', 'name': '石墨化温度5', 'format': '0.00', 'unit': '℃'},
    {'key': 'graphite.graphite_thickness5', 'name': '石墨化厚度5', 'format': '0.00', 'unit': 'μm'},
    {'key': 'graphite.graphite_temp6', 'name': '石墨化温度6', 'format': '0.00', 'unit': '℃'},
    {'key': 'graphite.graphite_thickness6', 'name': '石墨化厚度6', 'format': '0.00', 'unit': 'μm'},
    {'key': 'graphite.inner_foaming_thickness', 'name': '卷内发泡厚度', 'format': '0.00', 'unit': 'μm'},
    {'key': 'graphite.outer_foaming_thickness', 'name': '卷外发泡厚度', 'format': '0.00', 'unit': 'μm'},
    {'key': 'graphite.shrinkage_ratio', 'name': '收缩比', 'format': '0.00', 'unit': '%'},
    {'key': 'graphite.graphite_max_temp', 'name': '石墨化最高温度', 'format': '0.00', 'unit': '℃'},
    {'key': 'graphite.graphite_total_time', 'name': '石墨化总时长', 'format': '0', 'unit': 'min'},
    {'key': 'graphite.graphite_after_weight', 'name': '石墨化后重量', 'format': '0.000', 'unit': 'kg'},
    {'key': 'graphite.graphite_yield_rate', 'name': '石墨化收率', 'format': '0.00', 'unit': '%'},
    {'key': 'graphite.graphite_width', 'name': '石墨宽幅', 'format': '0.00', 'unit': 'mm'},
    {'key': 'graphite.gas_pressure', 'name': '气压值', 'format': '0.0000'},
    {'key': 'graphite.graphite_power', 'name': '石墨化电量', 'format': '0.00'},
    {'key': 'graphite.graphite_min_thickness', 'name': '石墨最薄极限', 'format': '0.00', 'unit': 'μm'},  # ✅ 新增
    {'key': 'graphite.graphite_notes', 'name': '石墨化补充说明', 'format': None},  # ✅ 新增
    
    # 压延参数（5个）
    {'key': 'rolling.rolling_machine', 'name': '压延机台', 'format': None},
    {'key': 'rolling.rolling_pressure', 'name': '压延压力', 'format': '0.00', 'unit': 'MPa'},
    {'key': 'rolling.rolling_tension', 'name': '压延张力', 'format': '0.00'},
    {'key': 'rolling.rolling_speed', 'name': '压延速度', 'format': '0.000', 'unit': 'm/s'},
    {'key': 'rolling.rolling_notes', 'name': '压延补充说明', 'format': None},  # ✅ 新增
    
    # 成品参数（13个）
    {'key': 'product.avg_thickness', 'name': '样品平均厚度', 'format': '0.00', 'unit': 'μm'},
    {'key': 'product.avg_density', 'name': '平均密度', 'format': '0.000', 'unit': 'g/cm³'},
    {'key': 'product.thermal_diffusivity', 'name': '热扩散系数', 'format': '0.000000', 'unit': 'mm²/s'},
    {'key': 'product.thermal_conductivity', 'name': '导热系数', 'format': '0.000', 'unit': 'W/m*K'},
    {'key': 'product.specific_heat', 'name': '比热', 'format': '0.0000', 'unit': 'J/g/K'},
    {'key': 'product.cohesion', 'name': '内聚力', 'format': '0.00', 'unit': 'gf'},
    {'key': 'product.peel_strength', 'name': '剥离力', 'format': '0.00', 'unit': 'gf'},
    {'key': 'product.bond_strength', 'name': '结合力', 'format': '0.00'},
    {'key': 'product.roughness', 'name': '粗糙度', 'format': None},
    {'key': 'product.appearance_desc', 'name': '外观描述', 'format': None},
    {'key': 'product.product_code', 'name': '成品编码', 'format': None},
    {'key': 'product.specification', 'name': '规格', 'format': None},
    {'key': 'product.experiment_summary', 'name': '实验总结', 'format': None},
]


def get_nested_value(obj: dict, key: str):
    """获取嵌套字典的值"""
    keys = key.split('.')
    value = obj
    for k in keys:
        if isinstance(value, dict):
            value = value.get(k)
        else:
            return None
        if value is None:
            return None
    return value


def convert_value_for_excel(value, number_format=None):
    """
    转换值为Excel友好格式
    
    关键点：
    1. Decimal → float（保留精度）
    2. None → '' （空字符串）
    3. 保持数值类型（不转字符串）
    """
    if value is None:
        return ''
    
    # Decimal转float（保留完整精度）
    if isinstance(value, Decimal):
        return float(value)
    
    # Date/Datetime转字符串
    if isinstance(value, (datetime, date)):
        return value.strftime('%Y-%m-%d') if isinstance(value, date) else value.strftime('%Y-%m-%d %H:%M:%S')
    
    # 其他类型直接返回
    return value


@compare_bp.route('/export', methods=['POST'])
@jwt_required()
def export_comparison():
    """
    导出实验对比数据为Excel文件
    
    请求体:
        {
            "experiment_ids": [1, 2, 3, 4]
        }
    
    返回:
        Excel文件流
    """
    try:
        # 1. 获取当前用户
        current_user_id = int(get_jwt_identity())
        user = User.query.get(current_user_id)
        
        if not user:
            return jsonify({'error': '用户不存在'}), 404
        
        # 2. 获取实验ID列表
        data = request.get_json()
        experiment_ids = data.get('experiment_ids', [])
        
        if not isinstance(experiment_ids, list) or len(experiment_ids) < 2:
            return jsonify({'error': '请选择至少2个实验进行对比'}), 400
        
        if len(experiment_ids) > 10:
            return jsonify({'error': '最多只能对比10个实验'}), 400
        
        print(f"\n{'='*60}")
        print(f"📤 导出实验对比数据")
        print(f"   用户: {user.username} (ID: {current_user_id})")
        print(f"   实验数量: {len(experiment_ids)}")
        print(f"   实验IDs: {experiment_ids}")
        print(f"{'='*60}\n")
        
        # 3. 查询实验数据（复用 compare_experiments 的逻辑）
        experiments = []
        for exp_id in experiment_ids:
            exp = Experiment.query.get(exp_id)
            if not exp:
                return jsonify({'error': f'实验 {exp_id} 不存在'}), 404
            
            # 权限检查
            if user.role == 'user' and exp.created_by != current_user_id:
                return jsonify({'error': '您没有权限查看此实验'}), 403
            
            # 获取完整数据
            exp_data = {
                'id': exp.id,
                'code': exp.experiment_code,
                'status': exp.status,
                'created_at': exp.created_at.strftime('%Y-%m-%d') if exp.created_at else '',
                'basic': {},
                'pi': {},
                'loose': {},
                'carbon': {},
                'graphite': {},
                'rolling': {},
                'product': {}
            }
            
            # 基本参数
            basic = ExperimentBasic.query.filter_by(experiment_id=exp.id).first()
            if basic:
                exp_data['basic'] = model_to_dict(basic)
            
            # PI膜参数
            pi = ExperimentPi.query.filter_by(experiment_id=exp.id).first()
            if pi:
                exp_data['pi'] = model_to_dict(pi)
            
            # 松卷参数
            loose = ExperimentLoose.query.filter_by(experiment_id=exp.id).first()
            if loose:
                exp_data['loose'] = model_to_dict(loose)
            
            # 碳化参数
            carbon = ExperimentCarbon.query.filter_by(experiment_id=exp.id).first()
            if carbon:
                exp_data['carbon'] = model_to_dict(carbon)
            
            # 石墨化参数
            graphite = ExperimentGraphite.query.filter_by(experiment_id=exp.id).first()
            if graphite:
                exp_data['graphite'] = model_to_dict(graphite)
            
            # 压延参数
            rolling = ExperimentRolling.query.filter_by(experiment_id=exp.id).first()
            if rolling:
                exp_data['rolling'] = model_to_dict(rolling)
            
            # 成品参数
            product = ExperimentProduct.query.filter_by(experiment_id=exp.id).first()
            if product:
                exp_data['product'] = model_to_dict(product)
            
            experiments.append(exp_data)
        
        print(f"✅ 数据查询完成，开始生成Excel...\n")
        
        # 4. 生成Excel文件（在内存中）
        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "实验对比数据"
        
        # 5. 写入表头（第1行：参数名称）
        ws['A1'] = '参数名称'
        ws['A1'].font = Font(bold=True, size=11)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        
        # 设置参数名列宽
        ws.column_dimensions['A'].width = 25
        
        # 写入实验编码表头
        for col_idx, exp in enumerate(experiments, start=2):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = exp['code']
            cell.font = Font(bold=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            
            # 设置数值列宽
            ws.column_dimensions[cell.column_letter].width = 15
        
        # 6. 写入数据行
        for row_idx, field in enumerate(EXPORT_FIELD_ORDER, start=2):
            # 参数名称（第1列）
            param_name = field['name']
            if field.get('unit'):
                param_name += f" ({field['unit']})"
            
            ws.cell(row=row_idx, column=1, value=param_name)
            ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row=row_idx, column=1).font = Font(size=10)
            
            # 实验数据（第2列开始）
            for col_idx, exp in enumerate(experiments, start=2):
                value = get_nested_value(exp, field['key'])
                converted_value = convert_value_for_excel(value, field.get('format'))
                
                cell = ws.cell(row=row_idx, column=col_idx, value=converted_value)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(size=10)
                
                # 设置数值格式（保留精度，显示美观）
                if field.get('format') and isinstance(converted_value, (int, float)):
                    cell.number_format = field['format']
        
        # 7. 添加边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=len(EXPORT_FIELD_ORDER) + 1, 
                                min_col=1, max_col=len(experiments) + 1):
            for cell in row:
                cell.border = thin_border
        
        # 8. 冻结首行和首列
        ws.freeze_panes = 'B2'
        
        # 9. 保存到内存流
        wb.save(output)
        output.seek(0)
        
        # 10. 生成文件名
        exp_count = len(experiments)
        today = datetime.now().strftime('%Y%m%d')
        filename = f'实验对比_{exp_count}个实验_{today}.xlsx'
        
        # URL编码文件名（防止中文乱码）
        encoded_filename = quote(filename)
        
        print(f"✅ Excel生成成功")
        print(f"   文件名: {filename}")
        print(f"   字段数: {len(EXPORT_FIELD_ORDER)}")
        print(f"   实验数: {len(experiments)}")
        print(f"{'='*60}\n")
        
        # 11. 返回文件流
        response = send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename  # Flask 2.0+ 推荐使用这个参数
        )
        
        # ✅ 关键：手动设置 Header，避免 TypeError 且支持中文文件名
        response.headers["Content-Disposition"] = f"attachment; filename*=UTF-8''{encoded_filename}"
        
        return response
        
    except Exception as e:
        print(f"\n❌ 导出失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'导出失败: {str(e)}'}), 500


# ============ 更新总结 ============
# 修复日期: 2024-12-09
# 修复内容: Decimal类型尾随零问题
# 
# 问题: thermal_diffusivity DECIMAL(10,6) 导致显示 999.000000
# 解决: model_to_dict() 中将 Decimal 转为 float，自动去除尾随零
# 
# 影响字段:
# - thermal_diffusivity (999.000000 → 999)
# - thermal_conductivity (1500.000 → 1500)
# - specific_heat (0.8500 → 0.85)
# - 所有 DECIMAL 类型字段
# 
# Excel导出功能添加日期: 2025-01-01
# 新增功能:
# - POST /api/compare/export - 导出Excel文件
# - EXPORT_FIELD_ORDER - 86个字段定义
# - 内存流处理（BytesIO）
# - 数值精度控制（number_format）
# - 中文文件名支持
# ====================================